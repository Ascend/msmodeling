#!/usr/bin/env python3
"""Layer Compare — 比较两个层 CSV 的 Stage 算子并排对比。

输出 xlsx 两个 sheet:
  Sheet1 "总比较": 按 Stage 汇总时间对比
  Sheet2 "算子明细": 逐算子并排对比

用法：
  python layer_compare.py -a forward_003_layer.csv -b kernel_layered_layer.csv
  python layer_compare.py -a forward_003_layer.csv -b kernel_layered_layer.csv -o compare.xlsx
"""

import argparse
import csv
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COMM_RE = re.compile(
    r"all_reduce|all_gather|allgather|reduce_scatter|reduceScatter|allReduce|hcom_",
    re.IGNORECASE,
)

# Named constants (replacing magic numbers)
SHAPE_TRUNCATE_LEN = 40
COL_WIDTH_PADDING = 4
SUMMARY_MAX_COL_WIDTH = 40
DETAIL_MAX_COL_WIDTH = 50
# Excel 公式注入防护：以这些字符开头的值需要转义
FORMULA_INJECTION_PREFIXES = ("=", "+", "-", "@")


def sanitize_cell(value):
    """转义可能触发 Excel 公式注入的值。"""
    if isinstance(value, str) and value.startswith(FORMULA_INJECTION_PREFIXES):
        return f"'{value}"
    return value


def load_layer_csv(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def simplify_shape(shape_str: str) -> str:
    """简化 shape 字符串，只保留维度数字。"""
    if not shape_str or shape_str.strip().upper() in ("", "N/A", "NULL"):
        return ""
    nums = re.findall(r"\d+", shape_str)
    return "x".join(nums) if nums else shape_str[:40]


def split_by_stages(
    rows: list[dict],
) -> tuple[list[tuple[str, list[dict]]], dict[str, float]]:
    """按 Stage 列分段，返回 ([(stage_display_name, [rows])], {stage: comm_duration})。

    所有 RSN 段统一用 "RSN"，不区分序号。
    通信算子排除，但记录被排除的通信耗时。
    """
    segments = []
    current_stage = None
    current_rows = []
    comm_by_stage: dict[str, float] = {}

    for r in rows:
        stage = r.get("Stage", "").strip()
        name = r.get("Name", "") + " " + r.get("Full Name", "")
        # 检测通信算子
        is_comm = bool(COMM_RE.search(name))
        dur = float(r.get("Duration(us)", 0))

        if is_comm:
            # 通信算子不计入 stage rows，但累加到对应 stage 的通信耗时
            target_stage = stage.strip() if stage.strip() else (current_stage or "")
            if target_stage:
                comm_by_stage[target_stage] = comm_by_stage.get(target_stage, 0.0) + dur
            continue

        if stage:
            if current_stage is not None and stage != current_stage:
                segments.append((current_stage, current_rows))
                current_stage = stage
                current_rows = [r]
            elif current_stage is None:
                current_stage = stage
                current_rows = [r]
            else:
                current_rows.append(r)
        elif current_stage is not None:
            current_rows.append(r)

    if current_stage is not None:
        segments.append((current_stage, current_rows))

    return segments, comm_by_stage


def compare(file_a: Path, file_b: Path, output: Path) -> None:
    rows_a = load_layer_csv(file_a)
    rows_b = load_layer_csv(file_b)

    segs_a, comm_a_map = split_by_stages(rows_a)
    segs_b, comm_b_map = split_by_stages(rows_b)

    name_a = file_a.stem
    name_b = file_b.stem

    # 按 Stage 对齐（保留原始顺序，RSN 不合并）
    max_segs = max(len(segs_a), len(segs_b))

    # ── 构建 Sheet1: 总比较 ──
    summary_rows = []
    for si in range(max_segs):
        stage = segs_a[si][0] if si < len(segs_a) else (segs_b[si][0] if si < len(segs_b) else "")
        ops_a = segs_a[si][1] if si < len(segs_a) else []
        ops_b = segs_b[si][1] if si < len(segs_b) else []
        dur_a = sum(float(r.get("Duration(us)", 0)) for r in ops_a)
        dur_b = sum(float(r.get("Duration(us)", 0)) for r in ops_b)
        comm_a = comm_a_map.get(stage, 0.0)
        comm_b = comm_b_map.get(stage, 0.0)
        diff = round(dur_a - dur_b, 3)
        pct = f"{diff / dur_b * 100:.2f}%" if dur_b else "N/A"
        summary_rows.append(
            {
                "Stage": stage,
                f"{name_a}_算子数": len(ops_a),
                f"{name_b}_算子数": len(ops_b),
                f"{name_a}_Duration(us)": round(dur_a, 3),
                f"{name_b}_Duration(us)": round(dur_b, 3),
                f"{name_a}_排除通信(us)": round(comm_a, 3),
                f"{name_b}_排除通信(us)": round(comm_b, 3),
                "Diff(us)": diff,
                "Diff(%)": pct,
            }
        )

    # 总计行
    tot_a = sum(r[f"{name_a}_Duration(us)"] for r in summary_rows)
    tot_b = sum(r[f"{name_b}_Duration(us)"] for r in summary_rows)
    tot_comm_a = sum(r[f"{name_a}_排除通信(us)"] for r in summary_rows)
    tot_comm_b = sum(r[f"{name_b}_排除通信(us)"] for r in summary_rows)
    tot_diff = round(tot_a - tot_b, 3)
    tot_pct = f"{tot_diff / tot_b * 100:.2f}%" if tot_b else "N/A"
    summary_rows.append(
        {
            "Stage": "TOTAL",
            f"{name_a}_算子数": sum(r[f"{name_a}_算子数"] for r in summary_rows),
            f"{name_b}_算子数": sum(r[f"{name_b}_算子数"] for r in summary_rows),
            f"{name_a}_Duration(us)": round(tot_a, 3),
            f"{name_b}_Duration(us)": round(tot_b, 3),
            f"{name_a}_排除通信(us)": round(tot_comm_a, 3),
            f"{name_b}_排除通信(us)": round(tot_comm_b, 3),
            "Diff(us)": tot_diff,
            "Diff(%)": tot_pct,
        }
    )

    # ── 构建 Sheet2: 算子明细（保留原始 Stage 顺序，RSN 不合并）──
    detail_rows = []
    max_segs = max(len(segs_a), len(segs_b))
    for si in range(max_segs):
        stage = segs_a[si][0] if si < len(segs_a) else (segs_b[si][0] if si < len(segs_b) else "")
        ops_a = segs_a[si][1] if si < len(segs_a) else []
        ops_b = segs_b[si][1] if si < len(segs_b) else []
        max_len = max(len(ops_a), len(ops_b))

        stage_total_a = 0.0
        stage_total_b = 0.0

        for i in range(max_len):
            ra = ops_a[i] if i < len(ops_a) else None
            rb = ops_b[i] if i < len(ops_b) else None

            na_val = ra.get("Name", "") if ra else ""
            nb_val = rb.get("Name", "") if rb else ""
            dur_a = float(ra.get("Duration(us)", 0)) if ra else 0.0
            dur_b = float(rb.get("Duration(us)", 0)) if rb else 0.0
            sha = simplify_shape(ra.get("Output Shapes", "")) if ra else ""
            shb = simplify_shape(rb.get("Output Shapes", "")) if rb else ""

            diff = round(dur_a - dur_b, 3) if (ra and rb) else ""
            stage_total_a += dur_a
            stage_total_b += dur_b

            detail_rows.append(
                {
                    "Stage": stage,
                    "序号": i + 1,
                    f"{name_a}_Name": na_val,
                    f"{name_a}_Duration(us)": round(dur_a, 3) if ra else "",
                    f"{name_a}_Shape": sha,
                    f"{name_b}_Name": nb_val,
                    f"{name_b}_Duration(us)": round(dur_b, 3) if rb else "",
                    f"{name_b}_Shape": shb,
                    "Duration_Diff(us)": "",
                }
            )

        # Stage 小计
        diff_total = round(stage_total_a - stage_total_b, 3)
        pct = f"{diff_total / stage_total_b * 100:.2f}%" if stage_total_b else "N/A"
        detail_rows.append(
            {
                "Stage": f"  {stage} 小计",
                "序号": "",
                f"{name_a}_Name": "",
                f"{name_a}_Duration(us)": round(stage_total_a, 3),
                f"{name_a}_Shape": "",
                f"{name_b}_Name": "",
                f"{name_b}_Duration(us)": round(stage_total_b, 3),
                f"{name_b}_Shape": "",
                "Duration_Diff(us)": f"{diff_total} ({pct})",
            }
        )

    # ── 写 xlsx ──
    wb = Workbook()

    # 样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    total_fill = PatternFill("solid", fgColor="FFC000")
    subtotal_fill = PatternFill("solid", fgColor="D9E2F3")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header(ws, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def style_rows(ws, data, ncols, is_summary=False):
        for ri, row_data in enumerate(data, start=2):
            is_total = row_data.get("Stage", "") == "TOTAL"
            is_subtotal = row_data.get("Stage", "").startswith("  ") and "小计" in row_data.get("Stage", "")
            for ci, key in enumerate(row_data.keys(), start=1):
                cell = ws.cell(row=ri, column=ci)
                cell.border = thin_border
                if is_total:
                    cell.fill = total_fill
                    cell.font = Font(bold=True)
                elif is_subtotal:
                    cell.fill = subtotal_fill

    # Sheet1: 总比较
    ws1 = wb.active
    ws1.title = "总比较"
    if summary_rows:
        headers = list(summary_rows[0].keys())
        ws1.append(headers)
        for row in summary_rows:
            ws1.append([sanitize_cell(row[h]) for h in headers])
        style_header(ws1, len(headers))
        style_rows(ws1, summary_rows, len(headers), is_summary=True)
        # 自动列宽
        for col_idx, h in enumerate(headers, 1):
            max_len = max(len(str(h)), *(len(str(row[h])) for row in summary_rows))
            col_letter = ws1.cell(row=1, column=col_idx).column_letter
            ws1.column_dimensions[col_letter].width = min(max_len + COL_WIDTH_PADDING, SUMMARY_MAX_COL_WIDTH)

    # Sheet2: 算子明细
    ws2 = wb.create_sheet("算子明细")
    if detail_rows:
        headers = list(detail_rows[0].keys())
        ws2.append(headers)
        for row in detail_rows:
            ws2.append([sanitize_cell(row[h]) for h in headers])
        style_header(ws2, len(headers))
        style_rows(ws2, detail_rows, len(headers))
        for col_idx, h in enumerate(headers, 1):
            max_len = max(len(str(h)), *(len(str(row[h])) for row in detail_rows))
            col_letter = ws2.cell(row=1, column=col_idx).column_letter
            ws2.column_dimensions[col_letter].width = min(max_len + COL_WIDTH_PADDING, DETAIL_MAX_COL_WIDTH)

    wb.save(str(output))
    print(f"\nXLSX: {output}")
    print(f"  Sheet1: 总比较 ({len(summary_rows)} rows)")
    print(f"  Sheet2: 算子明细 ({len(detail_rows)} rows)")

    # 终端打印总比较
    print(f"\n{'=' * 80}")
    print(f"Layer Compare: {name_a} vs {name_b}")
    print(f"{'=' * 80}")
    for r in summary_rows:
        stage = r["Stage"]
        da = r[f"{name_a}_Duration(us)"]
        db = r[f"{name_b}_Duration(us)"]
        diff = r["Diff(us)"]
        pct = r["Diff(%)"]
        print(f"  {stage:15s} | {da:>12} vs {db:>12} | diff={diff:>10} ({pct})")
    print(f"{'=' * 80}")


def main() -> int:
    parser = argparse.ArgumentParser(description="比较两个层 CSV 的 Stage 算子并排对比（输出 xlsx）")
    parser.add_argument("-a", required=True, help="文件A（层 CSV，来自 npu_layer_analyzer）")
    parser.add_argument("-b", required=True, help="文件B（层 CSV，来自 layer_analyzer）")
    parser.add_argument("-o", default=None, help="输出 xlsx 路径（默认 compare_result.xlsx）")
    args = parser.parse_args()

    file_a = Path(args.a)
    file_b = Path(args.b)
    output = Path(args.o) if args.o else file_a.parent / "compare_result.xlsx"

    if not file_a.is_file():
        raise SystemExit(f"文件不存在: {file_a}")
    if not file_b.is_file():
        raise SystemExit(f"文件不存在: {file_b}")

    compare(file_a, file_b, output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
